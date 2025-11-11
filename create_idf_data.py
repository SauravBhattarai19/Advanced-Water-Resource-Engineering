"""
Create IDF data Excel file for teaching purposes
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_idf_excel():
    """Create comprehensive IDF data Excel file"""

    # Generate realistic 60-minute annual maximum rainfall data
    np.random.seed(2024)  # For reproducible results
    years = list(range(1990, 2024))  # 34 years of data

    # Generate 60-minute rainfall depths (mm) - realistic for urban area
    base_mean = 40  # mm in 60 minutes
    rainfall_60min = []

    for year in years:
        # Add some variability and occasional extreme events
        if year in [1998, 2005, 2012, 2019]:  # extreme years
            value = np.random.gamma(shape=4, scale=base_mean/2) * 1.6
        else:
            value = np.random.gamma(shape=3, scale=base_mean/3)
        rainfall_60min.append(max(15, min(120, value)))  # Keep realistic bounds

    # Round to 1 decimal place
    rainfall_60min = np.round(rainfall_60min, 1)

    # Create main data DataFrame
    data = pd.DataFrame({
        'Year': years,
        'Rainfall_60min_mm': rainfall_60min
    })

    # Add intensity column
    data['Intensity_60min_mmhr'] = data['Rainfall_60min_mm']  # 60min data, so intensity = depth

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Raw Data
    ws1 = wb.create_sheet("60min_Annual_Maxima")

    # Add headers
    headers = ['Year', 'Rainfall (mm)', 'Intensity (mm/hr)', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Add data
    for row, (year, rainfall) in enumerate(zip(data['Year'], data['Rainfall_60min_mm']), 2):
        ws1.cell(row=row, column=1, value=year)
        ws1.cell(row=row, column=2, value=rainfall)
        ws1.cell(row=row, column=3, value=rainfall)  # For 60min, intensity = depth

        # Add notes for extreme years
        if year in [1998, 2005, 2012, 2019]:
            ws1.cell(row=row, column=4, value="Extreme event year")

    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws1.column_dimensions[column_letter].width = adjusted_width

    # Sheet 2: Statistics Summary
    ws2 = wb.create_sheet("Data_Statistics")

    # Calculate statistics
    stats_data = [
        ['Statistic', 'Value', 'Units'],
        ['Number of Years', len(data), 'years'],
        ['Minimum', data['Rainfall_60min_mm'].min(), 'mm'],
        ['Maximum', data['Rainfall_60min_mm'].max(), 'mm'],
        ['Mean', round(data['Rainfall_60min_mm'].mean(), 1), 'mm'],
        ['Median', round(data['Rainfall_60min_mm'].median(), 1), 'mm'],
        ['Standard Deviation', round(data['Rainfall_60min_mm'].std(), 1), 'mm'],
        ['Coefficient of Variation', round(data['Rainfall_60min_mm'].std()/data['Rainfall_60min_mm'].mean(), 2), '-'],
    ]

    for row, stat_row in enumerate(stats_data, 1):
        for col, value in enumerate(stat_row, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            if row == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    # Auto-adjust column widths
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws2.column_dimensions[column_letter].width = adjusted_width

    # Sheet 3: Instructions
    ws3 = wb.create_sheet("Instructions")

    instructions = [
        ['IDF Curve Analysis - 60-Minute Rainfall Data'],
        [''],
        ['PURPOSE:'],
        ['This dataset contains 34 years of annual maximum 60-minute rainfall data'],
        ['for use in creating Intensity-Duration-Frequency (IDF) curves.'],
        [''],
        ['DATA SOURCE:'],
        ['Synthetic data based on realistic rainfall patterns for educational purposes.'],
        ['Data represents a temperate climate region with occasional extreme events.'],
        [''],
        ['USAGE INSTRUCTIONS:'],
        ['1. Use the "60min_Annual_Maxima" sheet for your analysis'],
        ['2. Apply disaggregation ratios to create shorter duration data'],
        ['3. Perform frequency analysis for each duration'],
        ['4. Create IDF curves using the results'],
        [''],
        ['DISAGGREGATION RATIOS (typical values):'],
        ['5-min:   P5 = P60 × 0.25'],
        ['10-min:  P10 = P60 × 0.35'],
        ['15-min:  P15 = P60 × 0.45'],
        ['30-min:  P30 = P60 × 0.65'],
        ['60-min:  P60 = P60 × 1.00 (base data)'],
        ['120-min: P120 = P60 × 1.30'],
        [''],
        ['GOOGLE COLAB LINK:'],
        ['For distribution analysis and best-fit testing:'],
        ['https://colab.research.google.com/drive/1t-Sz6p3xeyxV74efFzu6_gFsigLAkbHz?usp=sharing'],
        [''],
        ['NEXT STEPS:'],
        ['1. Copy this data to the Google Colab notebook'],
        ['2. Run the distribution analysis to find the best-fit distribution'],
        ['3. Use the results to create accurate IDF curves'],
        ['4. Apply the curves for engineering design problems'],
    ]

    for row, instruction in enumerate(instructions, 1):
        cell = ws3.cell(row=row, column=1, value=instruction[0] if instruction else '')
        if row == 1:  # Title
            cell.font = Font(size=14, bold=True, color="366092")
        elif instruction and instruction[0].endswith(':'):  # Section headers
            cell.font = Font(bold=True, color="366092")
        elif instruction and instruction[0].startswith('https'):  # Link
            cell.font = Font(color="0000FF", underline="single")

    # Auto-adjust column width
    ws3.column_dimensions['A'].width = 80

    # Sheet 4: Excel Template
    ws4 = wb.create_sheet("Excel_Template")

    # Create template headers
    template_headers = [
        'Year', 'P_60min', 'P_5min', 'P_10min', 'P_15min', 'P_30min', 'P_120min',
        'I_5min', 'I_10min', 'I_15min', 'I_30min', 'I_60min', 'I_120min'
    ]

    for col, header in enumerate(template_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

    # Add sample formulas for first few rows
    for row in range(2, 6):  # First 4 data rows
        year_idx = row - 2
        if year_idx < len(data):
            # Year and base 60-min data
            ws4.cell(row=row, column=1, value=data.iloc[year_idx]['Year'])
            ws4.cell(row=row, column=2, value=data.iloc[year_idx]['Rainfall_60min_mm'])

            # Disaggregation formulas (showing formulas for first row, values for others)
            if row == 2:
                ws4.cell(row=row, column=3, value='=B2*0.25')  # 5-min
                ws4.cell(row=row, column=4, value='=B2*0.35')  # 10-min
                ws4.cell(row=row, column=5, value='=B2*0.45')  # 15-min
                ws4.cell(row=row, column=6, value='=B2*0.65')  # 30-min
                ws4.cell(row=row, column=7, value='=B2*1.30')  # 120-min

                ws4.cell(row=row, column=8, value='=(C2/5)*60')   # I_5min
                ws4.cell(row=row, column=9, value='=(D2/10)*60')  # I_10min
                ws4.cell(row=row, column=10, value='=(E2/15)*60') # I_15min
                ws4.cell(row=row, column=11, value='=(F2/30)*60') # I_30min
                ws4.cell(row=row, column=12, value='=B2')         # I_60min
                ws4.cell(row=row, column=13, value='=(G2/120)*60') # I_120min

    # Auto-adjust column widths
    for column in ws4.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws4.column_dimensions[column_letter].width = adjusted_width

    # Save the file
    output_path = "notebooks/idf.xlsx"
    wb.save(output_path)
    print(f"IDF data file created successfully: {output_path}")

    return output_path

if __name__ == "__main__":
    create_idf_excel()