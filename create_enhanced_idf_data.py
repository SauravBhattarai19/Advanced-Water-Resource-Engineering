"""
Create Enhanced IDF data Excel file with Year, Month, Day, Hour columns for pivot table functionality
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_enhanced_idf_excel():
    """Create comprehensive IDF data Excel file with date/time breakdown"""

    # Generate realistic 60-minute annual maximum rainfall data
    np.random.seed(2024)  # For reproducible results
    years = list(range(1990, 2024))  # 34 years of data

    # Generate data with realistic seasonal patterns
    data_records = []

    for year in years:
        # Generate a realistic date for annual maximum (more likely in summer months)
        # Weight towards months 5-9 (May through September) for temperate climate
        month_weights = [0.02, 0.02, 0.05, 0.10, 0.15, 0.20, 0.20, 0.15, 0.08, 0.02, 0.01, 0.01]
        month = np.random.choice(range(1, 13), p=month_weights)

        # Generate day within the month
        days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        if month == 2 and year % 4 == 0:  # Leap year
            days_in_month[1] = 29
        day = np.random.randint(1, days_in_month[month - 1] + 1)

        # Generate hour (afternoon storms more common)
        hour_weights = [0.02] * 6 + [0.03] * 6 + [0.08] * 6 + [0.12] * 6  # 6am-6pm more likely
        hour = np.random.choice(range(24), p=hour_weights)

        # Generate 60-minute rainfall depths (mm) - realistic for urban area
        base_mean = 40  # mm in 60 minutes

        # Add some variability and occasional extreme events
        if year in [1998, 2005, 2012, 2019]:  # extreme years
            rainfall = np.random.gamma(shape=4, scale=base_mean/2) * 1.6
        else:
            rainfall = np.random.gamma(shape=3, scale=base_mean/3)

        # Keep realistic bounds
        rainfall = max(15, min(120, rainfall))
        rainfall = round(rainfall, 1)

        # Create the date
        date = datetime(year, month, day, hour)

        data_records.append({
            'DateTime': date,
            'Year': year,
            'Month': month,
            'Day': day,
            'Hour': hour,
            'Rainfall_60min_mm': rainfall,
            'Intensity_60min_mmhr': rainfall  # For 60min, intensity = depth
        })

    # Create main data DataFrame
    data = pd.DataFrame(data_records)

    # Sort by year to maintain chronological order
    data = data.sort_values('Year').reset_index(drop=True)

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Raw Data
    ws1 = wb.create_sheet("60min_Annual_Maxima")

    # Add headers
    headers = ['Year', 'Month', 'Day', 'Hour', 'DateTime', 'Rainfall (mm)', 'Intensity (mm/hr)', 'Season', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Add data
    for row, record in enumerate(data.to_dict('records'), 2):
        ws1.cell(row=row, column=1, value=record['Year'])
        ws1.cell(row=row, column=2, value=record['Month'])
        ws1.cell(row=row, column=3, value=record['Day'])
        ws1.cell(row=row, column=4, value=record['Hour'])
        ws1.cell(row=row, column=5, value=record['DateTime'].strftime('%Y-%m-%d %H:%M'))
        ws1.cell(row=row, column=6, value=record['Rainfall_60min_mm'])
        ws1.cell(row=row, column=7, value=record['Rainfall_60min_mm'])  # For 60min, intensity = depth

        # Add season
        month = record['Month']
        if month in [12, 1, 2]:
            season = "Winter"
        elif month in [3, 4, 5]:
            season = "Spring"
        elif month in [6, 7, 8]:
            season = "Summer"
        else:
            season = "Fall"
        ws1.cell(row=row, column=8, value=season)

        # Add notes for extreme years
        if record['Year'] in [1998, 2005, 2012, 2019]:
            ws1.cell(row=row, column=9, value="Extreme event year")

    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws1.column_dimensions[column_letter].width = adjusted_width

    # Sheet 2: Statistics Summary
    ws2 = wb.create_sheet("Data_Statistics")

    # Calculate statistics
    stats_data = [
        ['Statistic', 'Value', 'Units'],
        ['Number of Years', len(data), 'years'],
        ['Date Range', f"{data['Year'].min()}-{data['Year'].max()}", ''],
        ['Minimum Rainfall', data['Rainfall_60min_mm'].min(), 'mm'],
        ['Maximum Rainfall', data['Rainfall_60min_mm'].max(), 'mm'],
        ['Mean Rainfall', round(data['Rainfall_60min_mm'].mean(), 1), 'mm'],
        ['Median Rainfall', round(data['Rainfall_60min_mm'].median(), 1), 'mm'],
        ['Standard Deviation', round(data['Rainfall_60min_mm'].std(), 1), 'mm'],
        ['Coefficient of Variation', round(data['Rainfall_60min_mm'].std()/data['Rainfall_60min_mm'].mean(), 2), '-'],
        ['', '', ''],
        ['SEASONAL DISTRIBUTION', '', ''],
        ['Spring Events', sum(data['Month'].isin([3,4,5])), 'count'],
        ['Summer Events', sum(data['Month'].isin([6,7,8])), 'count'],
        ['Fall Events', sum(data['Month'].isin([9,10,11])), 'count'],
        ['Winter Events', sum(data['Month'].isin([12,1,2])), 'count'],
    ]

    for row, stat_row in enumerate(stats_data, 1):
        for col, value in enumerate(stat_row, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            if row == 1 or stat_row[0] == 'SEASONAL DISTRIBUTION':  # Header rows
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    # Auto-adjust column widths
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws2.column_dimensions[column_letter].width = adjusted_width

    # Sheet 3: Pivot Table Guide
    ws3 = wb.create_sheet("Pivot_Table_Guide")

    pivot_instructions = [
        ['Pivot Table Analysis Guide'],
        [''],
        ['AVAILABLE FIELDS FOR PIVOT TABLES:'],
        ['• Year - Analyze trends over time'],
        ['• Month - Seasonal patterns (1=Jan, 12=Dec)'],
        ['• Day - Daily patterns within months'],
        ['• Hour - Hourly patterns (0-23)'],
        ['• Season - Seasonal grouping (Spring, Summer, Fall, Winter)'],
        ['• Rainfall (mm) - Values to summarize'],
        ['• Intensity (mm/hr) - Same as rainfall for 60-min data'],
        [''],
        ['SUGGESTED PIVOT TABLE ANALYSES:'],
        [''],
        ['1. SEASONAL ANALYSIS:'],
        ['   Rows: Season'],
        ['   Values: Average of Rainfall (mm)'],
        ['   Result: Compare seasonal rainfall patterns'],
        [''],
        ['2. MONTHLY TRENDS:'],
        ['   Rows: Month'],
        ['   Values: Average of Rainfall (mm), Count of Year'],
        ['   Result: Monthly distribution of extreme events'],
        [''],
        ['3. HOURLY PATTERNS:'],
        ['   Rows: Hour'],
        ['   Values: Average of Rainfall (mm)'],
        ['   Result: Time of day when extreme events occur'],
        [''],
        ['4. DECADAL ANALYSIS:'],
        ['   Rows: Year (grouped by decades)'],
        ['   Values: Average of Rainfall (mm), Maximum of Rainfall (mm)'],
        ['   Result: Climate change trends'],
        [''],
        ['5. COMBINED ANALYSIS:'],
        ['   Rows: Season, Month'],
        ['   Columns: Year (grouped by 5-year periods)'],
        ['   Values: Average of Rainfall (mm)'],
        ['   Result: Detailed temporal patterns'],
        [''],
        ['HOW TO CREATE PIVOT TABLES:'],
        ['1. Select all data in "60min_Annual_Maxima" sheet'],
        ['2. Insert > PivotTable'],
        ['3. Choose fields from the list above'],
        ['4. Drag fields to Rows, Columns, Values areas'],
        ['5. Analyze patterns and trends'],
    ]

    for row, instruction in enumerate(pivot_instructions, 1):
        cell = ws3.cell(row=row, column=1, value=instruction[0] if instruction else '')
        if row == 1:  # Title
            cell.font = Font(size=14, bold=True, color="366092")
        elif instruction and (instruction[0].endswith(':') or instruction[0].startswith('AVAILABLE') or instruction[0].startswith('SUGGESTED')):  # Section headers
            cell.font = Font(bold=True, color="366092")
        elif instruction and instruction[0].startswith('   '):  # Indented items
            cell.font = Font(color="444444")

    # Auto-adjust column width
    ws3.column_dimensions['A'].width = 80

    # Sheet 4: Instructions
    ws4 = wb.create_sheet("Instructions")

    instructions = [
        ['IDF Curve Analysis - Enhanced 60-Minute Rainfall Data'],
        [''],
        ['PURPOSE:'],
        ['This dataset contains 34 years of annual maximum 60-minute rainfall data'],
        ['with enhanced date/time breakdown for comprehensive analysis using pivot tables.'],
        [''],
        ['DATA STRUCTURE:'],
        ['• Year, Month, Day, Hour: Complete temporal breakdown'],
        ['• DateTime: Combined date and time stamp'],
        ['• Rainfall (mm): 60-minute precipitation depth'],
        ['• Intensity (mm/hr): Same as rainfall for 60-min duration'],
        ['• Season: Meteorological season classification'],
        ['• Notes: Special annotations for extreme events'],
        [''],
        ['DATA SOURCE:'],
        ['Synthetic data based on realistic rainfall patterns for educational purposes.'],
        ['Data represents a temperate climate region with seasonal variability.'],
        [''],
        ['USAGE INSTRUCTIONS:'],
        ['1. Use the "60min_Annual_Maxima" sheet for your analysis'],
        ['2. Create pivot tables to analyze temporal patterns'],
        ['3. Apply disaggregation ratios to create shorter duration data'],
        ['4. Perform frequency analysis for each duration'],
        ['5. Create IDF curves using the results'],
        [''],
        ['DISAGGREGATION RATIOS (typical values):'],
        ['5-min:   P5 = P60 × 0.25'],
        ['10-min:  P10 = P60 × 0.35'],
        ['15-min:  P15 = P60 × 0.45'],
        ['30-min:  P30 = P60 × 0.65'],
        ['60-min:  P60 = P60 × 1.00 (base data)'],
        ['120-min: P120 = P60 × 1.30'],
        [''],
        ['GOOGLE COLAB LINK:'],
        ['For distribution analysis and best-fit testing:'],
        ['https://colab.research.google.com/drive/1t-Sz6p3xeyxV74efFzu6_gFsigLAkbHz?usp=sharing'],
        [''],
        ['PIVOT TABLE WORKFLOW:'],
        ['1. Select data range in "60min_Annual_Maxima" sheet'],
        ['2. Insert > PivotTable'],
        ['3. Analyze seasonal, monthly, and hourly patterns'],
        ['4. Export results for IDF curve development'],
        [''],
        ['NEXT STEPS:'],
        ['1. Explore data patterns using pivot tables'],
        ['2. Copy data to the Google Colab notebook for distribution analysis'],
        ['3. Run statistical tests to find the best-fit distribution'],
        ['4. Create accurate IDF curves for engineering design'],
    ]

    for row, instruction in enumerate(instructions, 1):
        cell = ws4.cell(row=row, column=1, value=instruction[0] if instruction else '')
        if row == 1:  # Title
            cell.font = Font(size=14, bold=True, color="366092")
        elif instruction and instruction[0].endswith(':'):  # Section headers
            cell.font = Font(bold=True, color="366092")
        elif instruction and instruction[0].startswith('https'):  # Link
            cell.font = Font(color="0000FF", underline="single")

    # Auto-adjust column width
    ws4.column_dimensions['A'].width = 80

    # Sheet 5: Excel Template (Enhanced)
    ws5 = wb.create_sheet("Excel_Template")

    # Create template headers
    template_headers = [
        'Year', 'Month', 'Day', 'Hour', 'P_60min', 'P_5min', 'P_10min', 'P_15min', 'P_30min', 'P_120min',
        'I_5min', 'I_10min', 'I_15min', 'I_30min', 'I_60min', 'I_120min'
    ]

    for col, header in enumerate(template_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

    # Add sample formulas for first few rows
    for row in range(2, 6):  # First 4 data rows
        data_idx = row - 2
        if data_idx < len(data):
            # Date/time and base 60-min data
            record = data.iloc[data_idx]
            ws5.cell(row=row, column=1, value=record['Year'])
            ws5.cell(row=row, column=2, value=record['Month'])
            ws5.cell(row=row, column=3, value=record['Day'])
            ws5.cell(row=row, column=4, value=record['Hour'])
            ws5.cell(row=row, column=5, value=record['Rainfall_60min_mm'])

            # Disaggregation formulas (showing formulas for first row, values for others)
            if row == 2:
                ws5.cell(row=row, column=6, value='=E2*0.25')  # 5-min
                ws5.cell(row=row, column=7, value='=E2*0.35')  # 10-min
                ws5.cell(row=row, column=8, value='=E2*0.45')  # 15-min
                ws5.cell(row=row, column=9, value='=E2*0.65')  # 30-min
                ws5.cell(row=row, column=10, value='=E2*1.30')  # 120-min

                ws5.cell(row=row, column=11, value='=(F2/5)*60')   # I_5min
                ws5.cell(row=row, column=12, value='=(G2/10)*60')  # I_10min
                ws5.cell(row=row, column=13, value='=(H2/15)*60')  # I_15min
                ws5.cell(row=row, column=14, value='=(I2/30)*60')  # I_30min
                ws5.cell(row=row, column=15, value='=E2')          # I_60min
                ws5.cell(row=row, column=16, value='=(J2/120)*60') # I_120min

    # Auto-adjust column widths
    for column in ws5.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws5.column_dimensions[column_letter].width = adjusted_width

    # Save the file
    output_path = "notebooks/idf.xlsx"
    wb.save(output_path)
    print(f"Enhanced IDF data file created successfully: {output_path}")
    print(f"Total records: {len(data)}")
    print(f"Date range: {data['Year'].min()}-{data['Year'].max()}")
    print(f"Seasons represented: {sorted(data.groupby(data['Month'].apply(lambda x: {12:\"Winter\", 1:\"Winter\", 2:\"Winter\", 3:\"Spring\", 4:\"Spring\", 5:\"Spring\", 6:\"Summer\", 7:\"Summer\", 8:\"Summer\", 9:\"Fall\", 10:\"Fall\", 11:\"Fall\"}[x])).groups.keys())}")

    return output_path

if __name__ == "__main__":
    create_enhanced_idf_excel()